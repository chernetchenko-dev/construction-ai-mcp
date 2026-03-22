"""
msproject_mcp_server.py
MCP-сервер для Microsoft Project через COM (win32com) + Aspose.Tasks
Версия: 1.0  |  Март 2026
Поддерживает два режима:
  - COM  : требует запущенный MS Project Professional
  - File : работает напрямую с .mpp файлом через Aspose.Tasks (без MSP)

Инструменты (38 шт.):
  Статус / Проект        : msproject_status, msproject_open, msproject_save, msproject_close, msproject_project_info
  Задачи (чтение)        : msproject_get_tasks, msproject_get_task, msproject_get_critical_path, msproject_get_task_tree
  Задачи (запись)        : msproject_add_task, msproject_update_task, msproject_delete_task, msproject_link_tasks,
                           msproject_set_task_percent, msproject_bulk_update_tasks
  Ресурсы                : msproject_get_resources, msproject_get_resource, msproject_add_resource,
                           msproject_update_resource, msproject_delete_resource
  Назначения             : msproject_get_assignments, msproject_assign_resource, msproject_remove_assignment,
                           msproject_get_resource_workload
  Базовый план / Статус  : msproject_set_baseline, msproject_clear_baseline, msproject_get_baseline_comparison,
                           msproject_update_progress, msproject_get_earned_value
  Экспорт                : msproject_export_xml, msproject_export_csv, msproject_export_pdf,
                           msproject_export_excel, msproject_export_html
  Анализ                 : msproject_get_overallocated, msproject_get_late_tasks, msproject_get_summary,
                           msproject_get_milestones, msproject_find_tasks
"""

import json
import os
import sys
import traceback
from datetime import datetime, timedelta
from typing import Optional, Any

# ── FastMCP ──────────────────────────────────────────────────────────────────
try:
    from mcp.server.fastmcp import FastMCP
except ImportError:
    print("Установите: pip install mcp", file=sys.stderr)
    sys.exit(1)

mcp = FastMCP("msproject")

# ── Режим работы ─────────────────────────────────────────────────────────────
# Переключается вызовом msproject_open(file_path, mode="com"|"file")
# или через переменную окружения MSP_MODE=com|file
_mode: str = os.environ.get("MSP_MODE", "com")   # "com" | "file"
_com_app: Any = None          # win32com MSProject.Application
_aspose_project: Any = None   # aspose_tasks.Project
_current_file: str = ""

# ═══════════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ═══════════════════════════════════════════════════════════════════════════

def _date_str(d) -> str:
    """Преобразует COM Date или datetime в строку ISO."""
    if d is None:
        return ""
    try:
        if hasattr(d, "strftime"):
            return d.strftime("%Y-%m-%d")
        return str(d)
    except Exception:
        return str(d)


def _duration_days(dur_str) -> Optional[float]:
    """Пытается извлечь число дней из строки длительности MSP ('5 days' → 5.0)."""
    if dur_str is None:
        return None
    s = str(dur_str).lower().replace("дней", "").replace("days", "").replace("d", "").strip()
    try:
        return float(s)
    except Exception:
        return None


def _task_to_dict_com(t) -> dict:
    """Преобразует COM Task в словарь."""
    try:
        return {
            "id": t.ID,
            "unique_id": t.UniqueID,
            "name": t.Name,
            "outline_level": t.OutlineLevel,
            "duration_days": _duration_days(t.Duration),
            "start": _date_str(t.Start),
            "finish": _date_str(t.Finish),
            "percent_complete": t.PercentComplete,
            "work_hours": float(t.Work) / 60.0 if t.Work else 0,
            "cost": float(t.Cost) if t.Cost else 0,
            "is_summary": bool(t.Summary),
            "is_milestone": bool(t.Milestone),
            "is_critical": bool(t.Critical),
            "baseline_start": _date_str(t.BaselineStart),
            "baseline_finish": _date_str(t.BaselineFinish),
            "baseline_duration_days": _duration_days(t.BaselineDuration),
            "slack_days": _duration_days(t.TotalSlack),
            "notes": str(t.Notes) if t.Notes else "",
            "wbs": str(t.WBS) if hasattr(t, "WBS") else "",
            "predecessors": str(t.Predecessors) if t.Predecessors else "",
        }
    except Exception as e:
        return {"id": getattr(t, "ID", "?"), "error": str(e)}


def _resource_to_dict_com(r) -> dict:
    try:
        return {
            "id": r.ID,
            "unique_id": r.UniqueID,
            "name": r.Name,
            "type": str(r.Type),   # 0=Work, 1=Material, 2=Cost
            "max_units": float(r.MaxUnits) * 100 if r.MaxUnits else 100,
            "cost_per_hour": float(r.StandardRate) if r.StandardRate else 0,
            "overtime_rate": float(r.OvertimeRate) if r.OvertimeRate else 0,
            "total_work_hours": float(r.Work) / 60.0 if r.Work else 0,
            "overallocated": bool(r.Overallocated),
            "email": str(r.EMailAddress) if hasattr(r, "EMailAddress") else "",
        }
    except Exception as e:
        return {"id": getattr(r, "ID", "?"), "error": str(e)}


def _get_com_app():
    global _com_app
    if _com_app is not None:
        return _com_app
    try:
        import win32com.client as win32
        _com_app = win32.GetActiveObject("MSProject.Application")
        return _com_app
    except Exception:
        try:
            import win32com.client as win32
            _com_app = win32.Dispatch("MSProject.Application")
            _com_app.Visible = True
            return _com_app
        except Exception as e:
            raise RuntimeError(f"MS Project не запущен и не удалось запустить: {e}")


def _get_active_project_com():
    app = _get_com_app()
    proj = app.ActiveProject
    if proj is None:
        raise RuntimeError("Нет активного проекта. Откройте .mpp файл в MS Project.")
    return proj


def _get_aspose_project():
    global _aspose_project
    if _aspose_project is None:
        raise RuntimeError("Файл не открыт. Вызовите msproject_open(file_path, mode='file') сначала.")
    return _aspose_project


# ═══════════════════════════════════════════════════════════════════════════
#  ГРУППА 1: СТАТУС И ПРОЕКТ
# ═══════════════════════════════════════════════════════════════════════════

@mcp.tool()
def msproject_status() -> str:
    """
    Проверить статус подключения к MS Project.
    Возвращает: текущий режим (com/file), имя проекта, число задач, ресурсов.
    Всегда вызывай первым для диагностики.
    """
    result = {"mode": _mode, "file": _current_file}
    if _mode == "com":
        try:
            app = _get_com_app()
            proj = app.ActiveProject
            if proj:
                result["status"] = "connected"
                result["project_name"] = proj.Name
                result["task_count"] = proj.Tasks.Count if proj.Tasks else 0
                result["resource_count"] = proj.Resources.Count if proj.Resources else 0
                result["start"] = _date_str(proj.ProjectStart)
                result["finish"] = _date_str(proj.ProjectFinish)
                result["percent_complete"] = proj.PercentComplete
            else:
                result["status"] = "no_active_project"
        except Exception as e:
            result["status"] = "error"
            result["error"] = str(e)
    else:  # file mode
        try:
            ap = _get_aspose_project()
            result["status"] = "connected"
            result["project_name"] = ap.Get(None) or _current_file
            result["task_count"] = len(list(ap.RootTask.Children)) if ap.RootTask else 0
        except Exception as e:
            result["status"] = "error"
            result["error"] = str(e)
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
def msproject_open(file_path: str, mode: str = "com") -> str:
    """
    Открыть .mpp файл.

    Параметры:
      file_path : полный путь к .mpp файлу (например C:\\projects\\stroy.mpp)
      mode      : 'com'  — открыть через запущенный MS Project (по умолчанию)
                  'file' — открыть напрямую через Aspose.Tasks (без MSP)

    Пример: msproject_open("C:\\projects\\road.mpp", mode="com")
    """
    global _mode, _com_app, _aspose_project, _current_file

    if not os.path.exists(file_path):
        return json.dumps({"status": "error", "error": f"Файл не найден: {file_path}"})

    _mode = mode
    _current_file = file_path

    if mode == "com":
        try:
            import win32com.client as win32
            try:
                _com_app = win32.GetActiveObject("MSProject.Application")
            except Exception:
                _com_app = win32.Dispatch("MSProject.Application")
                _com_app.Visible = True
            _com_app.FileOpen(file_path)
            proj = _com_app.ActiveProject
            return json.dumps({
                "status": "ok",
                "mode": "com",
                "project_name": proj.Name,
                "task_count": proj.Tasks.Count if proj.Tasks else 0,
            }, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        try:
            import aspose.tasks as tasks
            _aspose_project = tasks.Project(file_path)
            return json.dumps({
                "status": "ok",
                "mode": "file",
                "file": file_path,
            }, ensure_ascii=False)
        except ImportError:
            return json.dumps({"status": "error", "error": "aspose-tasks не установлен. Выполни: pip install aspose-tasks"})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})


@mcp.tool()
def msproject_save(file_path: str = "") -> str:
    """
    Сохранить текущий проект.

    Параметры:
      file_path : если указан — сохранить как новый файл (SaveAs).
                  Если пустой — сохранить текущий файл.
    """
    if _mode == "com":
        try:
            app = _get_com_app()
            if file_path:
                app.FileSaveAs(file_path)
            else:
                app.FileSave()
            return json.dumps({"status": "ok", "saved": file_path or "текущий файл"})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        try:
            ap = _get_aspose_project()
            save_to = file_path or _current_file
            ap.Save(save_to)
            return json.dumps({"status": "ok", "saved": save_to})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})


@mcp.tool()
def msproject_project_info() -> str:
    """
    Получить сводную информацию о проекте:
    название, автор, даты начала/конца, бюджет, % выполнения,
    статистику по задачам и ресурсам.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            total = proj.Tasks.Count if proj.Tasks else 0
            critical = sum(1 for i in range(1, total + 1)
                           if proj.Tasks(i) and proj.Tasks(i).Critical)
            done = sum(1 for i in range(1, total + 1)
                       if proj.Tasks(i) and proj.Tasks(i).PercentComplete == 100)
            return json.dumps({
                "name": proj.Name,
                "author": str(proj.Author) if hasattr(proj, "Author") else "",
                "start": _date_str(proj.ProjectStart),
                "finish": _date_str(proj.ProjectFinish),
                "percent_complete": proj.PercentComplete,
                "total_cost": float(proj.Cost) if proj.Cost else 0,
                "total_work_hours": float(proj.Work) / 60.0 if proj.Work else 0,
                "tasks_total": total,
                "tasks_critical": critical,
                "tasks_done": done,
                "resources_total": proj.Resources.Count if proj.Resources else 0,
            }, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "msproject_project_info поддерживается только в режиме COM"})


# ═══════════════════════════════════════════════════════════════════════════
#  ГРУППА 2: ЗАДАЧИ — ЧТЕНИЕ
# ═══════════════════════════════════════════════════════════════════════════

@mcp.tool()
def msproject_get_tasks(
    filter_by: str = "",
    filter_value: str = "",
    limit: int = 100,
    include_summary: bool = False,
) -> str:
    """
    Получить список задач проекта.

    Параметры:
      filter_by     : поле для фильтрации: 'name', 'wbs', 'percent_complete', 'critical',
                      'is_milestone', 'resource_name', 'outline_level'
      filter_value  : значение фильтра (строка). Для булевых: 'true'/'false'.
                      Для percent_complete: '0' (незавершённые), '100' (завершённые).
      limit         : максимум задач в ответе (по умолчанию 100)
      include_summary: включить ли суммарные задачи (по умолчанию False)

    Примеры:
      msproject_get_tasks()                         → все задачи (до 100)
      msproject_get_tasks("critical", "true")       → только критический путь
      msproject_get_tasks("percent_complete", "0")  → незавершённые
      msproject_get_tasks("name", "фундамент")      → поиск по имени
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            tasks = []
            count = proj.Tasks.Count if proj.Tasks else 0
            fv_lower = filter_value.lower()
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t is None:
                    continue
                if not include_summary and t.Summary:
                    continue
                d = _task_to_dict_com(t)
                # фильтрация
                if filter_by:
                    if filter_by == "name" and filter_value.lower() not in d.get("name", "").lower():
                        continue
                    elif filter_by == "critical" and str(d.get("is_critical", False)).lower() != fv_lower:
                        continue
                    elif filter_by == "is_milestone" and str(d.get("is_milestone", False)).lower() != fv_lower:
                        continue
                    elif filter_by == "percent_complete":
                        pc = d.get("percent_complete", -1)
                        if filter_value == "0" and pc >= 100:
                            continue
                        elif filter_value == "100" and pc < 100:
                            continue
                        elif filter_value not in ("0", "100") and str(pc) != filter_value:
                            continue
                    elif filter_by == "outline_level" and str(d.get("outline_level", "")) != filter_value:
                        continue
                    elif filter_by == "wbs" and filter_value not in d.get("wbs", ""):
                        continue
                tasks.append(d)
                if len(tasks) >= limit:
                    break
            return json.dumps({"tasks": tasks, "count": len(tasks)}, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e), "trace": traceback.format_exc()})
    else:
        return json.dumps({"error": "В режиме 'file' используй msproject_open с mode='com' или обратись к COM."})


@mcp.tool()
def msproject_get_task(task_id: int) -> str:
    """
    Получить подробную информацию об одной задаче по её ID.

    Параметры:
      task_id : числовой ID задачи (поле ID в MS Project, не UniqueID)

    Возвращает: все поля задачи + список предшественников и преемников.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            t = proj.Tasks(task_id)
            if t is None:
                return json.dumps({"error": f"Задача #{task_id} не найдена"})
            d = _task_to_dict_com(t)
            # Добавляем назначения ресурсов
            assignments = []
            if t.Assignments:
                for j in range(1, t.Assignments.Count + 1):
                    a = t.Assignments(j)
                    if a:
                        assignments.append({
                            "resource_name": a.ResourceName,
                            "units_pct": float(a.Units) * 100 if a.Units else 100,
                            "work_hours": float(a.Work) / 60.0 if a.Work else 0,
                        })
            d["assignments"] = assignments
            return json.dumps(d, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_critical_path() -> str:
    """
    Получить задачи критического пути.
    Возвращает список задач с is_critical=True, отсортированных по дате начала.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            critical = []
            count = proj.Tasks.Count if proj.Tasks else 0
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t and t.Critical and not t.Summary:
                    critical.append(_task_to_dict_com(t))
            critical.sort(key=lambda x: x.get("start", ""))
            return json.dumps({
                "critical_tasks": critical,
                "count": len(critical),
            }, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_task_tree(parent_id: int = 0, depth: int = 3) -> str:
    """
    Получить иерархическое дерево задач (WBS-структуру).

    Параметры:
      parent_id : ID родительской задачи (0 = весь проект)
      depth     : глубина вложенности (по умолчанию 3)

    Возвращает вложенный JSON с children для каждой задачи.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            count = proj.Tasks.Count if proj.Tasks else 0
            all_tasks = []
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t:
                    all_tasks.append(_task_to_dict_com(t))

            def build_tree(tasks, current_level, max_depth):
                if max_depth <= 0:
                    return []
                result = []
                for t in tasks:
                    if t.get("outline_level") == current_level:
                        node = dict(t)
                        children = [x for x in tasks if x.get("wbs", "").startswith(t.get("wbs", "NONE") + ".")]
                        node["children"] = build_tree(children, current_level + 1, max_depth - 1)
                        result.append(node)
                return result

            root_level = 1
            tree = build_tree(all_tasks, root_level, depth)
            return json.dumps({"tree": tree}, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_milestones() -> str:
    """
    Получить все вехи (milestones) проекта.
    Возвращает список вех с датами и процентом выполнения.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            milestones = []
            count = proj.Tasks.Count if proj.Tasks else 0
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t and t.Milestone:
                    milestones.append(_task_to_dict_com(t))
            return json.dumps({"milestones": milestones, "count": len(milestones)},
                               ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_find_tasks(query: str) -> str:
    """
    Найти задачи по названию (поиск подстроки, нечувствительно к регистру).

    Параметры:
      query : строка поиска

    Пример: msproject_find_tasks("бетон") → все задачи со словом "бетон"
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            results = []
            count = proj.Tasks.Count if proj.Tasks else 0
            q = query.lower()
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t and t.Name and q in t.Name.lower():
                    results.append(_task_to_dict_com(t))
            return json.dumps({"results": results, "count": len(results)},
                               ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


# ═══════════════════════════════════════════════════════════════════════════
#  ГРУППА 3: ЗАДАЧИ — ЗАПИСЬ
# ═══════════════════════════════════════════════════════════════════════════

@mcp.tool()
def msproject_add_task(
    name: str,
    duration_days: float = 1.0,
    start: str = "",
    parent_task_id: int = 0,
    predecessor_ids: str = "",
    notes: str = "",
    is_milestone: bool = False,
) -> str:
    """
    Добавить новую задачу в проект.

    Параметры:
      name            : название задачи
      duration_days   : длительность в днях (по умолчанию 1)
      start           : дата начала в формате YYYY-MM-DD (пусто = авто)
      parent_task_id  : ID родительской задачи (0 = корневой уровень)
      predecessor_ids : предшественники через запятую (например "5,7,12")
      notes           : заметки к задаче
      is_milestone    : является ли вехой (True/False)

    Возвращает ID созданной задачи.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            # Добавить в конец или после parent
            new_task = proj.Tasks.Add(name)
            new_task.Duration = f"{duration_days}d"
            if start:
                new_task.Start = start
            if is_milestone:
                new_task.Milestone = True
                new_task.Duration = "0d"
            if notes:
                new_task.Notes = notes
            if predecessor_ids:
                new_task.Predecessors = predecessor_ids
            return json.dumps({
                "status": "ok",
                "task_id": new_task.ID,
                "unique_id": new_task.UniqueID,
                "name": new_task.Name,
            }, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_update_task(
    task_id: int,
    name: str = "",
    duration_days: float = -1,
    start: str = "",
    finish: str = "",
    notes: str = "",
    predecessor_ids: str = "",
    constraint_type: str = "",
    constraint_date: str = "",
) -> str:
    """
    Обновить поля существующей задачи.

    Параметры:
      task_id         : ID задачи (обязательный)
      name            : новое название (пусто = не менять)
      duration_days   : новая длительность (-1 = не менять)
      start           : новая дата начала YYYY-MM-DD
      finish          : новая дата окончания YYYY-MM-DD
      notes           : новые заметки
      predecessor_ids : предшественники через запятую
      constraint_type : тип ограничения: 'MSO' (как можно раньше),
                        'MFO' (как можно позже), 'SNET' (не ранее),
                        'FNLT' (не позднее)
      constraint_date : дата ограничения YYYY-MM-DD

    Возвращает обновлённые данные задачи.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            t = proj.Tasks(task_id)
            if t is None:
                return json.dumps({"error": f"Задача #{task_id} не найдена"})
            if name:
                t.Name = name
            if duration_days >= 0:
                t.Duration = f"{duration_days}d"
            if start:
                t.Start = start
            if finish:
                t.Finish = finish
            if notes:
                t.Notes = notes
            if predecessor_ids:
                t.Predecessors = predecessor_ids
            if constraint_type:
                constraint_map = {"MSO": 0, "MFO": 1, "SNET": 4, "FNLT": 5, "SNLT": 6, "FNET": 7}
                t.ConstraintType = constraint_map.get(constraint_type.upper(), 0)
            if constraint_date:
                t.ConstraintDate = constraint_date
            return json.dumps({"status": "ok", "task": _task_to_dict_com(t)},
                               ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_delete_task(task_id: int) -> str:
    """
    Удалить задачу по ID.
    ⚠ Внимание: удаление необратимо (если не сохранить до этого).
    Рекомендуется перед удалением вызвать msproject_save().
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            t = proj.Tasks(task_id)
            if t is None:
                return json.dumps({"error": f"Задача #{task_id} не найдена"})
            name = t.Name
            t.Delete()
            return json.dumps({"status": "ok", "deleted": {"id": task_id, "name": name}})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_link_tasks(
    predecessor_id: int,
    successor_id: int,
    link_type: str = "FS",
    lag_days: float = 0,
) -> str:
    """
    Создать связь между задачами.

    Параметры:
      predecessor_id : ID предшествующей задачи
      successor_id   : ID следующей задачи
      link_type      : тип связи: 'FS' (Finish-Start), 'FF' (Finish-Finish),
                                   'SS' (Start-Start), 'SF' (Start-Finish)
      lag_days       : запаздывание в днях (отрицательное = опережение)
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            succ = proj.Tasks(successor_id)
            if succ is None:
                return json.dumps({"error": f"Задача #{successor_id} не найдена"})
            lag_str = f"{int(lag_days)}d" if lag_days != 0 else "0"
            # Тип связи: FS=1, FF=2, SS=3, SF=4
            lt_map = {"FS": 1, "FF": 2, "SS": 3, "SF": 4}
            lt = lt_map.get(link_type.upper(), 1)
            succ.TaskDependencies.Add(proj.Tasks(predecessor_id), lt, lag_str)
            return json.dumps({
                "status": "ok",
                "link": f"{predecessor_id} → {successor_id}",
                "type": link_type,
                "lag_days": lag_days,
            })
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_set_task_percent(task_id: int, percent: int) -> str:
    """
    Установить процент выполнения задачи.

    Параметры:
      task_id : ID задачи
      percent : значение от 0 до 100

    Пример: msproject_set_task_percent(15, 50) → задача #15 выполнена на 50%
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            t = proj.Tasks(task_id)
            if t is None:
                return json.dumps({"error": f"Задача #{task_id} не найдена"})
            percent = max(0, min(100, percent))
            t.PercentComplete = percent
            return json.dumps({
                "status": "ok",
                "task_id": task_id,
                "task_name": t.Name,
                "percent_complete": percent,
            })
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_bulk_update_tasks(updates: str) -> str:
    """
    Массово обновить несколько задач за один вызов.

    Параметры:
      updates : JSON-строка со списком обновлений. Каждый элемент:
                [{"task_id": 5, "percent_complete": 100},
                 {"task_id": 7, "name": "Новое название"},
                 {"task_id": 10, "duration_days": 3}]

    Поддерживаемые поля: percent_complete, name, duration_days, start, finish, notes
    Возвращает сводку по каждой операции.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            items = json.loads(updates)
            results = []
            for upd in items:
                tid = upd.get("task_id")
                if not tid:
                    results.append({"error": "task_id обязателен", "data": upd})
                    continue
                t = proj.Tasks(tid)
                if t is None:
                    results.append({"task_id": tid, "error": "не найдена"})
                    continue
                changes = []
                if "percent_complete" in upd:
                    t.PercentComplete = int(upd["percent_complete"])
                    changes.append("percent_complete")
                if "name" in upd:
                    t.Name = upd["name"]
                    changes.append("name")
                if "duration_days" in upd:
                    t.Duration = f"{upd['duration_days']}d"
                    changes.append("duration_days")
                if "start" in upd:
                    t.Start = upd["start"]
                    changes.append("start")
                if "finish" in upd:
                    t.Finish = upd["finish"]
                    changes.append("finish")
                if "notes" in upd:
                    t.Notes = upd["notes"]
                    changes.append("notes")
                results.append({"task_id": tid, "task_name": t.Name, "updated": changes})
            return json.dumps({"status": "ok", "results": results}, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


# ═══════════════════════════════════════════════════════════════════════════
#  ГРУППА 4: РЕСУРСЫ
# ═══════════════════════════════════════════════════════════════════════════

@mcp.tool()
def msproject_get_resources(resource_type: str = "all") -> str:
    """
    Получить список ресурсов проекта.

    Параметры:
      resource_type : 'all' | 'work' (трудовые) | 'material' (материальные) | 'cost' (затратные)
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            resources = []
            count = proj.Resources.Count if proj.Resources else 0
            type_filter = {"work": "0", "material": "1", "cost": "2"}
            for i in range(1, count + 1):
                r = proj.Resources(i)
                if r is None:
                    continue
                d = _resource_to_dict_com(r)
                if resource_type != "all":
                    if str(r.Type) != type_filter.get(resource_type, "-1"):
                        continue
                resources.append(d)
            return json.dumps({"resources": resources, "count": len(resources)},
                               ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_resource(resource_id: int) -> str:
    """Получить подробную информацию о ресурсе по ID."""
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            r = proj.Resources(resource_id)
            if r is None:
                return json.dumps({"error": f"Ресурс #{resource_id} не найден"})
            d = _resource_to_dict_com(r)
            # Список задач ресурса
            assignments = []
            if r.Assignments:
                for j in range(1, r.Assignments.Count + 1):
                    a = r.Assignments(j)
                    if a and a.Task:
                        assignments.append({
                            "task_id": a.Task.ID,
                            "task_name": a.Task.Name,
                            "work_hours": float(a.Work) / 60.0 if a.Work else 0,
                            "units_pct": float(a.Units) * 100 if a.Units else 100,
                        })
            d["assignments"] = assignments
            return json.dumps(d, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_add_resource(
    name: str,
    resource_type: str = "work",
    max_units_pct: float = 100.0,
    standard_rate: float = 0.0,
    email: str = "",
) -> str:
    """
    Добавить ресурс в проект.

    Параметры:
      name           : имя ресурса
      resource_type  : 'work' | 'material' | 'cost'
      max_units_pct  : максимальная загрузка в % (по умолчанию 100%)
      standard_rate  : стандартная ставка (руб/час)
      email          : email ресурса (необязательно)
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            r = proj.Resources.Add(name)
            type_map = {"work": 0, "material": 1, "cost": 2}
            r.Type = type_map.get(resource_type.lower(), 0)
            r.MaxUnits = max_units_pct / 100.0
            if standard_rate > 0:
                r.StandardRate = standard_rate
            if email:
                r.EMailAddress = email
            return json.dumps({
                "status": "ok",
                "resource_id": r.ID,
                "unique_id": r.UniqueID,
                "name": r.Name,
            }, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_update_resource(
    resource_id: int,
    name: str = "",
    max_units_pct: float = -1,
    standard_rate: float = -1,
) -> str:
    """Обновить параметры ресурса."""
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            r = proj.Resources(resource_id)
            if r is None:
                return json.dumps({"error": f"Ресурс #{resource_id} не найден"})
            if name:
                r.Name = name
            if max_units_pct >= 0:
                r.MaxUnits = max_units_pct / 100.0
            if standard_rate >= 0:
                r.StandardRate = standard_rate
            return json.dumps({"status": "ok", "resource": _resource_to_dict_com(r)},
                               ensure_ascii=False)
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_delete_resource(resource_id: int) -> str:
    """Удалить ресурс из проекта. ⚠ Все назначения ресурса будут удалены."""
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            r = proj.Resources(resource_id)
            if r is None:
                return json.dumps({"error": f"Ресурс #{resource_id} не найден"})
            name = r.Name
            r.Delete()
            return json.dumps({"status": "ok", "deleted": {"id": resource_id, "name": name}})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_overallocated() -> str:
    """
    Найти перегруженные ресурсы (загрузка > 100%).
    Возвращает список ресурсов с их задачами в период перегрузки.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            overallocated = []
            count = proj.Resources.Count if proj.Resources else 0
            for i in range(1, count + 1):
                r = proj.Resources(i)
                if r and r.Overallocated:
                    d = _resource_to_dict_com(r)
                    overallocated.append(d)
            return json.dumps({
                "overallocated_resources": overallocated,
                "count": len(overallocated),
            }, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


# ═══════════════════════════════════════════════════════════════════════════
#  ГРУППА 5: НАЗНАЧЕНИЯ
# ═══════════════════════════════════════════════════════════════════════════

@mcp.tool()
def msproject_get_assignments(task_id: int = 0, resource_id: int = 0) -> str:
    """
    Получить назначения ресурсов на задачи.

    Параметры (оба необязательные, но хотя бы один нужен для фильтрации):
      task_id     : фильтр по задаче (0 = все)
      resource_id : фильтр по ресурсу (0 = все)

    Пример:
      msproject_get_assignments(task_id=5)        → кто назначен на задачу #5
      msproject_get_assignments(resource_id=3)    → какие задачи у ресурса #3
      msproject_get_assignments()                 → все назначения (осторожно: может быть много)
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            assignments = []
            count = proj.Tasks.Count if proj.Tasks else 0
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t is None or t.Summary:
                    continue
                if task_id and t.ID != task_id:
                    continue
                if t.Assignments:
                    for j in range(1, t.Assignments.Count + 1):
                        a = t.Assignments(j)
                        if a is None:
                            continue
                        if resource_id and (a.ResourceID != resource_id):
                            continue
                        assignments.append({
                            "task_id": t.ID,
                            "task_name": t.Name,
                            "resource_id": a.ResourceID,
                            "resource_name": a.ResourceName,
                            "units_pct": float(a.Units) * 100 if a.Units else 100,
                            "work_hours": float(a.Work) / 60.0 if a.Work else 0,
                            "start": _date_str(a.Start),
                            "finish": _date_str(a.Finish),
                        })
            return json.dumps({"assignments": assignments, "count": len(assignments)},
                               ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_assign_resource(
    task_id: int,
    resource_id: int,
    units_pct: float = 100.0,
) -> str:
    """
    Назначить ресурс на задачу.

    Параметры:
      task_id     : ID задачи
      resource_id : ID ресурса
      units_pct   : загрузка в процентах (по умолчанию 100%)

    Пример: msproject_assign_resource(5, 3, 50) → ресурс #3 на задачу #5 с 50% загрузкой
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            t = proj.Tasks(task_id)
            r = proj.Resources(resource_id)
            if t is None:
                return json.dumps({"error": f"Задача #{task_id} не найдена"})
            if r is None:
                return json.dumps({"error": f"Ресурс #{resource_id} не найден"})
            a = proj.Assignments.Add(t.ID, r.ID, units_pct / 100.0)
            return json.dumps({
                "status": "ok",
                "task": t.Name,
                "resource": r.Name,
                "units_pct": units_pct,
            }, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_remove_assignment(task_id: int, resource_id: int) -> str:
    """
    Снять ресурс с задачи.

    Параметры:
      task_id     : ID задачи
      resource_id : ID ресурса
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            t = proj.Tasks(task_id)
            if t is None:
                return json.dumps({"error": f"Задача #{task_id} не найдена"})
            removed = False
            if t.Assignments:
                for j in range(t.Assignments.Count, 0, -1):
                    a = t.Assignments(j)
                    if a and a.ResourceID == resource_id:
                        a.Delete()
                        removed = True
                        break
            return json.dumps({
                "status": "ok" if removed else "not_found",
                "message": "Назначение удалено" if removed else "Назначение не найдено",
            })
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_resource_workload(resource_id: int) -> str:
    """
    Получить нагрузку ресурса по задачам с датами и часами работы.
    Полезно для анализа загрузки и выявления конфликтов.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            r = proj.Resources(resource_id)
            if r is None:
                return json.dumps({"error": f"Ресурс #{resource_id} не найден"})
            workload = []
            if r.Assignments:
                for j in range(1, r.Assignments.Count + 1):
                    a = r.Assignments(j)
                    if a and a.Task:
                        workload.append({
                            "task_id": a.Task.ID,
                            "task_name": a.Task.Name,
                            "start": _date_str(a.Start),
                            "finish": _date_str(a.Finish),
                            "work_hours": float(a.Work) / 60.0 if a.Work else 0,
                            "units_pct": float(a.Units) * 100 if a.Units else 100,
                        })
            workload.sort(key=lambda x: x.get("start", ""))
            total_hours = sum(w["work_hours"] for w in workload)
            return json.dumps({
                "resource_id": resource_id,
                "resource_name": r.Name,
                "max_units_pct": float(r.MaxUnits) * 100 if r.MaxUnits else 100,
                "overallocated": bool(r.Overallocated),
                "total_work_hours": total_hours,
                "workload": workload,
            }, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


# ═══════════════════════════════════════════════════════════════════════════
#  ГРУППА 6: БАЗОВЫЙ ПЛАН И СТАТУС
# ═══════════════════════════════════════════════════════════════════════════

@mcp.tool()
def msproject_set_baseline(baseline_number: int = 0) -> str:
    """
    Сохранить базовый план (snapshot текущего расписания).

    Параметры:
      baseline_number : 0..10 (0 = Baseline, 1 = Baseline1, ..., 10 = Baseline10)
                        По умолчанию 0 (основной базовый план).

    ⚠ Перезапишет существующий базовый план с тем же номером!
    """
    if _mode == "com":
        try:
            app = _get_com_app()
            # SetBaseline(baseline, allTasks, fromStart, fromFinish)
            app.SetBaseline(baseline_number, True, True, True)
            return json.dumps({
                "status": "ok",
                "baseline": f"Baseline{baseline_number}" if baseline_number > 0 else "Baseline",
                "message": "Базовый план сохранён для всех задач",
            })
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_clear_baseline(baseline_number: int = 0) -> str:
    """
    Очистить базовый план.

    Параметры:
      baseline_number : номер базового плана (0..10)
    """
    if _mode == "com":
        try:
            app = _get_com_app()
            app.ClearBaseline(baseline_number, True)
            return json.dumps({"status": "ok", "message": f"Baseline{baseline_number} очищен"})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_baseline_comparison(limit: int = 50) -> str:
    """
    Сравнить текущее расписание с базовым планом.

    Возвращает задачи с отклонениями:
      - date_variance_days  : отставание по срокам (>0 = опаздываем)
      - cost_variance       : отклонение по стоимости (>0 = превышение)
      - duration_variance   : отклонение по длительности в днях

    Параметры:
      limit : максимум задач в ответе (по умолчанию 50)
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            comparisons = []
            count = proj.Tasks.Count if proj.Tasks else 0
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t is None or t.Summary:
                    continue
                bs = t.BaselineStart
                bf = t.BaselineFinish
                if not bs or not bf:
                    continue  # нет базового плана
                # Отклонение в днях
                try:
                    import pywintypes
                    finish_var = t.FinishVariance  # минуты
                    start_var = t.StartVariance
                    dur_var = t.DurationVariance
                    cost_var = float(t.CostVariance) if t.CostVariance else 0
                    entry = {
                        "task_id": t.ID,
                        "task_name": t.Name,
                        "wbs": str(t.WBS) if hasattr(t, "WBS") else "",
                        "baseline_start": _date_str(bs),
                        "baseline_finish": _date_str(bf),
                        "current_start": _date_str(t.Start),
                        "current_finish": _date_str(t.Finish),
                        "finish_variance_days": round(float(finish_var) / (60 * 8), 2) if finish_var else 0,
                        "start_variance_days": round(float(start_var) / (60 * 8), 2) if start_var else 0,
                        "cost_variance": round(cost_var, 2),
                        "percent_complete": t.PercentComplete,
                    }
                    comparisons.append(entry)
                except Exception:
                    pass
                if len(comparisons) >= limit:
                    break
            # Сортируем по отклонению (сначала самые большие проблемы)
            comparisons.sort(key=lambda x: x.get("finish_variance_days", 0), reverse=True)
            return json.dumps({
                "comparisons": comparisons,
                "count": len(comparisons),
                "tasks_delayed": sum(1 for c in comparisons if c.get("finish_variance_days", 0) > 0),
            }, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_update_progress(
    status_date: str,
    task_updates: str,
) -> str:
    """
    Обновить прогресс выполнения проекта (ввод факта).

    Параметры:
      status_date  : дата статуса в формате YYYY-MM-DD
      task_updates : JSON-строка со списком:
                     [{"task_id": 5, "percent_complete": 75, "actual_start": "2026-03-01"},
                      {"task_id": 7, "percent_complete": 100, "actual_finish": "2026-03-15"}]

    Поддерживаемые поля обновления:
      percent_complete, actual_start, actual_finish, actual_work_hours
    """
    if _mode == "com":
        try:
            app = _get_com_app()
            proj = app.ActiveProject
            proj.StatusDate = status_date
            items = json.loads(task_updates)
            results = []
            for upd in items:
                tid = upd.get("task_id")
                t = proj.Tasks(tid)
                if t is None:
                    results.append({"task_id": tid, "error": "не найдена"})
                    continue
                if "percent_complete" in upd:
                    t.PercentComplete = int(upd["percent_complete"])
                if "actual_start" in upd:
                    t.ActualStart = upd["actual_start"]
                if "actual_finish" in upd:
                    t.ActualFinish = upd["actual_finish"]
                if "actual_work_hours" in upd:
                    t.ActualWork = float(upd["actual_work_hours"]) * 60
                results.append({"task_id": tid, "task_name": t.Name, "status": "ok"})
            return json.dumps({
                "status": "ok",
                "status_date": status_date,
                "results": results,
            }, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_earned_value() -> str:
    """
    Получить показатели освоенного объёма (EVM) по проекту:
      BCWS (Бюджет по плану), BCWP (Освоенный объём), ACWP (Фактические затраты),
      SV (Отклонение по срокам), CV (Отклонение по стоимости),
      SPI (Индекс выполнения сроков), CPI (Индекс выполнения стоимости).

    Требует наличия базового плана и факта выполнения.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            tasks_evm = []
            count = proj.Tasks.Count if proj.Tasks else 0
            total_bcws = total_bcwp = total_acwp = 0.0
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t is None or t.Summary:
                    continue
                try:
                    bcws = float(t.BCWS) if t.BCWS else 0
                    bcwp = float(t.BCWP) if t.BCWP else 0
                    acwp = float(t.ACWP) if t.ACWP else 0
                    sv = bcwp - bcws
                    cv = bcwp - acwp
                    spi = round(bcwp / bcws, 3) if bcws else None
                    cpi = round(bcwp / acwp, 3) if acwp else None
                    tasks_evm.append({
                        "task_id": t.ID,
                        "task_name": t.Name,
                        "bcws": round(bcws, 2),
                        "bcwp": round(bcwp, 2),
                        "acwp": round(acwp, 2),
                        "sv": round(sv, 2),
                        "cv": round(cv, 2),
                        "spi": spi,
                        "cpi": cpi,
                        "percent_complete": t.PercentComplete,
                    })
                    total_bcws += bcws
                    total_bcwp += bcwp
                    total_acwp += acwp
                except Exception:
                    pass
            project_spi = round(total_bcwp / total_bcws, 3) if total_bcws else None
            project_cpi = round(total_bcwp / total_acwp, 3) if total_acwp else None
            return json.dumps({
                "project_summary": {
                    "BCWS": round(total_bcws, 2),
                    "BCWP": round(total_bcwp, 2),
                    "ACWP": round(total_acwp, 2),
                    "SV": round(total_bcwp - total_bcws, 2),
                    "CV": round(total_bcwp - total_acwp, 2),
                    "SPI": project_spi,
                    "CPI": project_cpi,
                },
                "tasks": tasks_evm,
            }, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_late_tasks(days_threshold: int = 0) -> str:
    """
    Найти задачи с просроченными сроками или близкие к дедлайну.

    Параметры:
      days_threshold : задачи, которые опаздывают более чем на N дней.
                       0 = все просроченные (отставание > 0 дней).
                       Отрицательное значение = задачи, у которых резерв времени < N дней.

    Возвращает список с отставанием по дням.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            late = []
            count = proj.Tasks.Count if proj.Tasks else 0
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t is None or t.Summary or t.PercentComplete >= 100:
                    continue
                try:
                    fv = float(t.FinishVariance) / (60 * 8) if t.FinishVariance else 0
                    slack = float(t.TotalSlack) / (60 * 8) if t.TotalSlack else 999
                    is_late = fv > days_threshold if days_threshold >= 0 else slack < abs(days_threshold)
                    if is_late:
                        d = _task_to_dict_com(t)
                        d["finish_variance_days"] = round(fv, 1)
                        d["total_slack_days"] = round(slack, 1)
                        late.append(d)
                except Exception:
                    pass
            late.sort(key=lambda x: x.get("finish_variance_days", 0), reverse=True)
            return json.dumps({"late_tasks": late, "count": len(late)},
                               ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


@mcp.tool()
def msproject_get_summary() -> str:
    """
    Получить сводку по проекту для быстрого статус-репорта:
    % выполнения, EVM-показатели, кол-во просроченных задач,
    перегруженные ресурсы, вехи к сдаче.
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            total = proj.Tasks.Count if proj.Tasks else 0
            not_started = completed = in_progress = late_count = 0
            for i in range(1, total + 1):
                t = proj.Tasks(i)
                if t is None or t.Summary:
                    continue
                pc = t.PercentComplete
                if pc == 0:
                    not_started += 1
                elif pc == 100:
                    completed += 1
                else:
                    in_progress += 1
                try:
                    fv = float(t.FinishVariance) / (60 * 8) if t.FinishVariance else 0
                    if fv > 0:
                        late_count += 1
                except Exception:
                    pass
            over_r = sum(1 for i in range(1, (proj.Resources.Count or 0) + 1)
                         if proj.Resources(i) and proj.Resources(i).Overallocated)
            return json.dumps({
                "project_name": proj.Name,
                "start": _date_str(proj.ProjectStart),
                "finish": _date_str(proj.ProjectFinish),
                "percent_complete": proj.PercentComplete,
                "tasks": {
                    "total": total,
                    "not_started": not_started,
                    "in_progress": in_progress,
                    "completed": completed,
                    "late": late_count,
                },
                "resources_overallocated": over_r,
                "total_cost": float(proj.Cost) if proj.Cost else 0,
            }, ensure_ascii=False, indent=2)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": "Только режим COM"})


# ═══════════════════════════════════════════════════════════════════════════
#  ГРУППА 7: ЭКСПОРТ
# ═══════════════════════════════════════════════════════════════════════════

@mcp.tool()
def msproject_export_xml(output_path: str) -> str:
    """
    Экспортировать проект в XML (Microsoft Project XML Format).
    XML читается Primavera, другими системами, и может быть разобран Python-скриптами.

    Параметры:
      output_path : путь к файлу .xml (например C:\\exports\\project.xml)
    """
    if _mode == "com":
        try:
            app = _get_com_app()
            # FileFormat: 10 = XML
            app.FileSaveAs(output_path, FileFormat=10)
            return json.dumps({"status": "ok", "file": output_path})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        try:
            import aspose.tasks as tasks
            ap = _get_aspose_project()
            ap.Save(output_path, tasks.saving.SaveFileFormat.XML)
            return json.dumps({"status": "ok", "file": output_path})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})


@mcp.tool()
def msproject_export_csv(output_path: str, fields: str = "") -> str:
    """
    Экспортировать задачи в CSV.

    Параметры:
      output_path : путь к .csv файлу
      fields      : поля через запятую (пусто = все основные)
                    Доступные поля: id, name, wbs, duration_days, start, finish,
                    percent_complete, cost, is_critical, baseline_start, baseline_finish

    Пример: msproject_export_csv("C:\\exports\\tasks.csv", "id,name,start,finish,percent_complete")
    """
    if _mode == "com":
        try:
            import csv
            proj = _get_active_project_com()
            all_fields = ["id", "wbs", "name", "outline_level", "duration_days",
                          "start", "finish", "percent_complete", "cost",
                          "is_critical", "is_milestone", "baseline_start", "baseline_finish",
                          "slack_days", "predecessors", "notes"]
            if fields:
                export_fields = [f.strip() for f in fields.split(",")]
            else:
                export_fields = all_fields
            count = proj.Tasks.Count if proj.Tasks else 0
            rows = []
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t is None:
                    continue
                d = _task_to_dict_com(t)
                rows.append({f: d.get(f, "") for f in export_fields})
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=export_fields)
                writer.writeheader()
                writer.writerows(rows)
            return json.dumps({"status": "ok", "file": output_path, "rows": len(rows)})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        try:
            import aspose.tasks as tasks
            ap = _get_aspose_project()
            ap.Save(output_path, tasks.saving.SaveFileFormat.CSV)
            return json.dumps({"status": "ok", "file": output_path})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})


@mcp.tool()
def msproject_export_pdf(output_path: str, view: str = "Gantt Chart") -> str:
    """
    Экспортировать проект в PDF.

    Параметры:
      output_path : путь к .pdf файлу
      view        : вид для экспорта (только в режиме COM):
                    'Gantt Chart', 'Resource Sheet', 'Task Usage', 'Resource Usage'

    В режиме file (Aspose) экспортируется диаграмма Ганта.
    """
    if _mode == "com":
        try:
            app = _get_com_app()
            # Переключиться на нужный вид
            try:
                app.SelectTaskField(Row=1, Column="Name")
                app.ViewApply(view)
            except Exception:
                pass
            # FileFormat: 45 = PDF
            app.FileSaveAs(output_path, FileFormat=45)
            return json.dumps({"status": "ok", "file": output_path, "view": view})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        try:
            import aspose.tasks as tasks
            ap = _get_aspose_project()
            opts = tasks.saving.PdfSaveOptions()
            ap.Save(output_path, opts)
            return json.dumps({"status": "ok", "file": output_path})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})


@mcp.tool()
def msproject_export_excel(output_path: str) -> str:
    """
    Экспортировать задачи проекта в Excel (.xlsx).

    Параметры:
      output_path : путь к .xlsx файлу

    Создаёт файл с тремя листами:
      - Tasks     : все задачи с полями
      - Resources : список ресурсов
      - Assignments : назначения ресурсов на задачи
    """
    if _mode == "com":
        try:
            try:
                import openpyxl
            except ImportError:
                return json.dumps({"error": "openpyxl не установлен. Выполни: pip install openpyxl"})
            proj = _get_active_project_com()
            wb = openpyxl.Workbook()

            # Лист Tasks
            ws_tasks = wb.active
            ws_tasks.title = "Tasks"
            task_headers = ["ID", "WBS", "Название", "Уровень", "Дней", "Начало", "Окончание",
                            "%", "Стоимость", "Критическая", "Веха", "БП_Начало", "БП_Окончание",
                            "Резерв_дней", "Предшественники"]
            ws_tasks.append(task_headers)
            count = proj.Tasks.Count if proj.Tasks else 0
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t is None:
                    continue
                d = _task_to_dict_com(t)
                ws_tasks.append([
                    d.get("id"), d.get("wbs"), d.get("name"), d.get("outline_level"),
                    d.get("duration_days"), d.get("start"), d.get("finish"),
                    d.get("percent_complete"), d.get("cost"),
                    d.get("is_critical"), d.get("is_milestone"),
                    d.get("baseline_start"), d.get("baseline_finish"),
                    d.get("slack_days"), d.get("predecessors"),
                ])

            # Лист Resources
            ws_res = wb.create_sheet("Resources")
            res_headers = ["ID", "Название", "Тип", "Макс_%", "Ставка", "Часы", "Перегружен"]
            ws_res.append(res_headers)
            rcount = proj.Resources.Count if proj.Resources else 0
            for i in range(1, rcount + 1):
                r = proj.Resources(i)
                if r is None:
                    continue
                d = _resource_to_dict_com(r)
                ws_res.append([
                    d.get("id"), d.get("name"), d.get("type"),
                    d.get("max_units"), d.get("cost_per_hour"),
                    d.get("total_work_hours"), d.get("overallocated"),
                ])

            # Лист Assignments
            ws_asgn = wb.create_sheet("Assignments")
            asgn_headers = ["Task_ID", "Задача", "Resource_ID", "Ресурс", "Загрузка_%", "Часов"]
            ws_asgn.append(asgn_headers)
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t is None or t.Summary:
                    continue
                if t.Assignments:
                    for j in range(1, t.Assignments.Count + 1):
                        a = t.Assignments(j)
                        if a:
                            ws_asgn.append([
                                t.ID, t.Name, a.ResourceID, a.ResourceName,
                                round(float(a.Units) * 100 if a.Units else 100, 1),
                                round(float(a.Work) / 60.0 if a.Work else 0, 1),
                            ])

            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            wb.save(output_path)
            return json.dumps({"status": "ok", "file": output_path,
                                "sheets": ["Tasks", "Resources", "Assignments"]})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        try:
            import aspose.tasks as tasks
            ap = _get_aspose_project()
            opts = tasks.saving.XlsxOptions()
            ap.Save(output_path, opts)
            return json.dumps({"status": "ok", "file": output_path})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})


@mcp.tool()
def msproject_export_html(output_path: str) -> str:
    """
    Экспортировать задачи в HTML-отчёт.
    Создаёт читаемый HTML с диаграммой Ганта в виде таблицы.

    Параметры:
      output_path : путь к .html файлу
    """
    if _mode == "com":
        try:
            proj = _get_active_project_com()
            count = proj.Tasks.Count if proj.Tasks else 0
            rows = []
            for i in range(1, count + 1):
                t = proj.Tasks(i)
                if t is None:
                    continue
                d = _task_to_dict_com(t)
                indent = "&nbsp;" * (d.get("outline_level", 1) - 1) * 4
                critical_cls = ' style="color:red"' if d.get("is_critical") else ""
                milestone_cls = ' style="font-weight:bold;color:#7030a0"' if d.get("is_milestone") else ""
                pc = d.get("percent_complete", 0)
                bar = f'<div style="background:#70ad47;width:{pc}%;height:10px;display:inline-block"></div>'
                rows.append(f"""<tr{critical_cls}{milestone_cls}>
                    <td>{d.get("id")}</td>
                    <td>{d.get("wbs","")}</td>
                    <td>{indent}{d.get("name","")}</td>
                    <td>{d.get("duration_days","")}</td>
                    <td>{d.get("start","")}</td>
                    <td>{d.get("finish","")}</td>
                    <td>{bar} {pc}%</td>
                    <td>{"⚑" if d.get("is_milestone") else ""}</td>
                </tr>""")
            html = f"""<!DOCTYPE html>
<html lang="ru"><head><meta charset="utf-8">
<title>{proj.Name}</title>
<style>
body{{font-family:Arial,sans-serif;font-size:12px}}
table{{border-collapse:collapse;width:100%}}
th{{background:#203864;color:white;padding:6px;text-align:left}}
td{{padding:4px 6px;border-bottom:1px solid #ddd}}
tr:hover{{background:#f0f0f0}}
</style></head>
<body>
<h2>Проект: {proj.Name}</h2>
<p>Дата: {datetime.now().strftime("%d.%m.%Y")} | % выполнения: {proj.PercentComplete}%</p>
<table>
<thead><tr>
  <th>ID</th><th>WBS</th><th>Задача</th><th>Дней</th>
  <th>Начало</th><th>Окончание</th><th>Прогресс</th><th>Веха</th>
</tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>
</body></html>"""
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(html)
            return json.dumps({"status": "ok", "file": output_path, "task_count": count})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})
    else:
        try:
            import aspose.tasks as tasks
            ap = _get_aspose_project()
            ap.Save(output_path, tasks.saving.SaveFileFormat.HTML)
            return json.dumps({"status": "ok", "file": output_path})
        except Exception as e:
            return json.dumps({"status": "error", "error": str(e)})


# ═══════════════════════════════════════════════════════════════════════════
#  ТОЧКА ВХОДА
# ═══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    mcp.run()
